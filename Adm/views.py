import csv
from django.forms import inlineformset_factory
from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from Void.models import Evento, Produto, ProdutoTamanho, country
from django.views.decorators.http import require_POST
from django.contrib import messages
import openpyxl
from django.core.files.base import ContentFile
from .forms import EventoForm, ProdutoForm, ProdutoTamanhoForm,UploadExcelForm, paisesForm 
from django.contrib.auth import authenticate, login,logout
from django.urls import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.core.paginator import Paginator
from django.db.models import Sum


def lista_produtos_tamanhos(request):
    
    produtos_list = Produto.objects.all()
    tamanhos_list = ProdutoTamanho.objects.all()
    for produto in produtos_list:
        tamanhos_sum = ProdutoTamanho.objects.filter(produto_id=produto).aggregate(total=Sum('stock_por_tamanho'))
        produto.stock = tamanhos_sum['total'] if tamanhos_sum['total'] else produto.stock
        
    paginator_produtos = Paginator(produtos_list, 10)  # Mostra 10 produtos por página
    page_number_produtos = request.GET.get('page')
    page_obj_produtos = paginator_produtos.get_page(page_number_produtos)

    if request.method == 'POST':
        
        form = UploadExcelForm(request.POST, request.FILES)
        
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(filename=ContentFile(excel_file.read()))
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                produto = Produto(
                    nome=row[0],
                    descricao=row[1],
                    preco=row[2],
                    stock=row[3],
                    sku=row[4],
                    categoria=row[5],
                    cor=row[6],
                    tamanho=row[7],
                )

                imagem_url = row[8] if len(row) > 9 else None# Ajuste o índice conforme a sua planilha
                if imagem_url:
                    try:
                        response = request.get(imagem_url)
                        img_temp = ContentFile(response.content)
                        produto.imagem.save(f"{row[5]}_image.jpg", img_temp)
                    except Exception as e:
                        print(f"Erro ao baixar a imagem de {imagem_url}: {e}")

                produto.save()

            return redirect('lista_produtos')
    else:
        form = UploadExcelForm()
    return render(request, 'Produto/lista_produtos.html', {'page_obj': page_obj_produtos,'form':form,'tamanhos': tamanhos_list})



def adicionar_produto(request):
    if request.method == 'POST':
        form = ProdutoForm(request.POST, request.FILES)  # Inclua request.FILES aqui
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Produto adicionado com sucesso!')
            return HttpResponseRedirect(reverse('lista_produtos'))
    else:
        form = ProdutoForm()
    return render(request, 'Produto/form_produto.html', {'form': form})



def editar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == 'POST':
        form = ProdutoForm(request.POST, request.FILES, instance=produto)  # Inclua request.FILES aqui
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Produto editado com sucesso!')
            return redirect('lista_produtos')
    else:
        form = ProdutoForm(instance=produto)
    return render(request, 'Produto/form_produto.html', {'form': form})

def eliminar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == "POST":
        produto.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'Produto eliminado com sucesso!')
        return redirect('lista_produtos')
    return render(request, 'Produto/confirmar_eliminar.html', {'produto': produto})

@require_POST
def eliminar_selecionados(request):
    # Obtém a lista de IDs dos produtos selecionados
    produtos_ids = request.POST.getlist('produto_id')
    
    # Elimina os produtos selecionados
    Produto.objects.filter(id__in=produtos_ids).delete()
    
    # Redireciona para a página de listagem de produtos
    return redirect('lista_produtos')

from django.shortcuts import redirect, render
from .forms import ProdutoTamanhoForm

def adicionar_tamanho(request):
    if request.method == 'POST':
        form = ProdutoTamanhoForm(request.POST)
        if form.is_valid():
            tamanho = form.save(commit=False)
            tamanho.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Tamanho adicionado com sucesso!')
            return HttpResponseRedirect(reverse('lista_produtos'))
    else:
        form = ProdutoTamanhoForm()
    return render(request, 'Produto/tamanhoForm.html', {'form': form})

def editar_tamanho(request, id):
    tamanho = get_object_or_404(ProdutoTamanho, id=id)
    if request.method == "POST":
        form = ProdutoTamanhoForm(request.POST, instance=tamanho)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Tamanho editado com sucesso!')
            return HttpResponseRedirect(reverse('lista_produtos'))
    else:
        form = ProdutoTamanhoForm(instance=tamanho)
    
    return render(request, 'Produto/editar_tamanho.html', {'form': form})

def excluir_tamanho(request, id):
    tamanho = get_object_or_404(ProdutoTamanho, id=id)
    if request.method == "POST":
        tamanho.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'Tamanho eliminado com sucesso!')
        return HttpResponseRedirect(reverse('lista_produtos'))
    
    return render(request, 'Produto/confirmar_exclusao_tamanho.html', {'tamanho': tamanho})

def login_view(request):
    # Se o user já está logado, redirecioná-lo.
    if request.user.is_authenticated:
        return redirect('admin')
    if request.method == 'POST':
        # Pegar as informações do formulário.
        username = request.POST.get('email')
        password = request.POST.get('password')

        # Autenticar o usuário.
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request,"Login bem sucedido") 
            return HttpResponseRedirect(reverse('admin'))
        else:
            # Se a autenticação falhou, retorne algum erro.
            messages.error(request,'Email ou password incorretos.')
            return HttpResponseRedirect(reverse('login'))
        # Método GET, renderize o formulário de login.
    return render(request, 'login.html')
    

def adm(request):
    
    return render(request,'adm.html')

def logout_view(request):
    logout(request)
    # Adiciona uma mensagem de sucesso
    messages.success(request, 'Logout efetuado com sucesso!')
    return HttpResponseRedirect(reverse('home'))

def exportar_produtos_csv(request):
    # Criar uma resposta do tipo arquivo
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="produtos.csv"'

    # Criar um escritor CSV usando a resposta como arquivo
    writer = csv.writer(response,delimiter=';')

    # Escrever o cabeçalho do CSV
    writer.writerow(['ID', 'Nome', 'Descrição', 'Preço', 'Estoque', 'SKU', 'Categoria', 'Cor', 'Tamanho'])

    # Obter os produtos da base de dados
    produtos = Produto.objects.all()

    # Escrever os dados de cada produto
    for produto in produtos:
        writer.writerow([produto.id, produto.nome, produto.descricao, produto.preco, produto.stock, produto.sku, produto.categoria, produto.cor, produto.tamanho])

    # Retornar a resposta que agora contém o arquivo CSV
    return response

def paises_adicionar(request):
    if request.method == 'POST':
        form = paisesForm(request.POST)  # Inclua request.FILES aqui
        if form.is_valid():
            form.save()
           # Adiciona uma mensagem de sucesso
            messages.success(request, 'País adicionado com sucesso!')
            return HttpResponseRedirect(reverse('listar_paises'))
    else:
        form = paisesForm()
    return render(request, 'Paises/paisesForm.html', {'form': form})

def listar_paises(request):
    paises = country.objects.all()
    if request.method == 'POST':
        excel_file = request.FILES["excel_file"]
        
        # Usa a biblioteca openpyxl para abrir o arquivo Excel
        wb = openpyxl.load_workbook(excel_file)
        # Seleciona a primeira planilha do arquivo
        worksheet = wb["Sheet1"]
        
        # Itera sobre as linhas da planilha
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            _, created = country.objects.update_or_create(
                name=row[0],
                acronimo=row[1],
                shipping=row[2],
            )
        return HttpResponseRedirect(reverse('listar_paises'))
    return render(request, 'Paises/paisesLista.html', {'paises': paises})

def editar_pais(request, id):
    pais = get_object_or_404(country, id=id)
    if request.method == "POST":
        form = paisesForm(request.POST, instance=pais)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'País editado com sucesso!')
            return HttpResponseRedirect(reverse('listar_paises'))
    else:
        form = paisesForm(instance=pais)
    return render(request, 'Paises/paisesForm.html', {'form': form})

# Exclui país
def excluir_pais(request, id):
    pais = get_object_or_404(country, id=id)
    if request.method == "POST":
        pais.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'País eliminado com sucesso!')
        return HttpResponseRedirect(reverse('listar_paises'))
    return render(request, 'Paises/excluirPais.html', {'pais': pais})




def lista_eventos(request):
    eventos = Evento.objects.all()
    return render(request, 'Adm/eventos/lista_eventos.html', {'eventos': eventos})

def evento_add(request):
    if request.method == "POST":
        form = EventoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Ritual adicionado com sucesso!')
            return HttpResponseRedirect(reverse('lista_eventos'))
    else:
        form = EventoForm()
    return render(request, 'Eventos/form.html', {'form': form})

def evento_edit(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        form = EventoForm(request.POST, request.FILES, instance=evento)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Ritual editado com sucesso!')
            return HttpResponseRedirect(reverse('lista_eventos'))
    else:
        form = EventoForm(instance=evento)
    return render(request, 'Eventos/form.html', {'form': form})

def evento_delete(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        evento.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'Ritual eliminado com sucesso!')
        return HttpResponseRedirect(reverse('lista_eventos'))
    return render(request, 'Eventos/delete.html', {'evento': evento})

def exportar_eventos_para_csv(request):
    # Cria a resposta com o tipo de conteúdo correto
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="eventos.csv"'

    # Cria um objeto writer para escrever no arquivo CSV
    writer = csv.writer(response)

    # Escreve o cabeçalho do CSV
    writer.writerow(['ID', 'Título', 'Data', 'Localização', 'Hiperligação', 'Caminho da Foto'])

    # Escreve as linhas de dados dos eventos
    eventos = Evento.objects.all()
    for evento in eventos:
        writer.writerow([
            evento.id, 
            evento.titulo, 
            evento.data, 
            evento.localizacao, 
            evento.hiperligacao, 
            request.build_absolute_uri(evento.foto.url) if evento.foto else None,
        ])

    return response