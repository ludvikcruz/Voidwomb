from decimal import Decimal
import json
from django.shortcuts import get_object_or_404, render,redirect
from django.urls import reverse
from django.contrib import messages
from payment.carrinho import add_to_cart
from payment.models import ItemDoCarrinho, Pagamento
from .models import Evento, Produto, ProdutoTamanho, country
from django.core.mail import send_mail
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from paypalrestsdk import Payment
import re
from django.http import JsonResponse

from django.db.models import F



def index(request):
    return render(request,'index.html')


def contact(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        subject = request.POST.get('subject')
        message = request.POST.get('body')

        # Construa a mensagem de e-mail
        full_message = f"Recebido de: {email}\n\n{message}"

        # Enviar e-mail
        send_mail(
            subject=subject,
            message=full_message,
            from_email=email,
            recipient_list=['voidwomb.band@gmail.com'],  # Substitua pelo e-mail que receberá a mensagem
        )
        
        return redirect('contact')
    return render(request,'contact.html')

def rituals(request):
    eventos_list = Evento.objects.all()
    paginator = Paginator(eventos_list, 4)  # Mostra 10 eventos por página

    page_number = request.GET.get('page')  # Pega o número da página do query string
    eventos = paginator.get_page(page_number)

    return render(request, 'rituals.html', {'eventos': eventos})

def store(request):
    produtos_list = Produto.objects.filter(stock__gt=0).order_by('-data_criacao')
    paginator = Paginator(produtos_list, 4)
    page_number = request.GET.get('page')
    produtos = paginator.get_page(page_number)
    
    return render(request,'store.html',{'produtos':produtos})



def about(request):
    return render(request,'about.html')

def adicionar_ao_carrinho(request, produto_id):
    cart = request.session.get('carrinho', {})
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Verificar se há estoque suficiente
    if produto.stock >= 1:
        produto.stock = F('stock') - 1
        produto.save()
        
        product_id_str = str(produto_id)
        if product_id_str in cart:
            if isinstance(cart[product_id_str], dict):
                cart[product_id_str]['quantidade'] += 1
            else:
                cart[product_id_str] = {
                    'quantidade': cart[product_id_str] + 1, 
                    'tamanho': 'unico',
                    'nome': produto.nome,
                    'size': 'unico',
                    'preco': str(produto.preco),
                    'sku': produto.sku,
                    'categoria': produto.categoria
                }
        else:
            cart[product_id_str] = {
                'quantidade': 1, 
                'tamanho': 'unico',
                'nome': produto.nome,
                'size': 'unico',
                'preco': str(produto.preco),
                'sku': produto.sku,
                'categoria': produto.categoria
            }
    else:
        # Exibir mensagem de erro informando que o produto está esgotado
        messages.error(request, f"O produto {produto.nome} está esgotado.")
        return HttpResponseRedirect(reverse('store'))
        
    request.session['carrinho'] = cart
    print(cart)
    return HttpResponseRedirect(reverse('store'))



def adicionar_dentro_carrinho(request, produto_id):
    cart = request.session.get('carrinho', {})
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Verificar se há estoque suficiente
    if produto.stock >= 1:
        produto.stock = F('stock_disponivel') - 1
        produto.save()
        
        product_id_str = str(produto_id)
        if product_id_str in cart:
            if isinstance(cart[product_id_str], dict):
                cart[product_id_str]['quantidade'] += 1
            else:
                cart[product_id_str] = {
                    'quantidade': cart[product_id_str] + 1, 
                    'tamanho': 'unico',
                    'nome': produto.nome,
                    'size': 'unico',
                    'preco': str(produto.preco),
                    'sku': produto.sku,
                    'categoria': produto.categoria
                }
        else:
            cart[product_id_str] = {
                'quantidade': 1, 
                'tamanho': 'unico',
                'nome': produto.nome,
                'size': 'unico',
                'preco': str(produto.preco),
                'sku': produto.sku,
                'categoria': produto.categoria
            }
    else:
        # Exibir mensagem de erro informando que o produto está esgotado
        messages.error(request, f"O produto {produto.nome} está esgotado.")
        return HttpResponseRedirect(reverse('carrinho'))
        
    request.session['carrinho'] = cart
    return HttpResponseRedirect(reverse('carrinho'))



def adicionar_roupa(request, produto_id):
     # Obtém o carrinho da sessão
    cart = request.session.get('carrinho', {})
    
    # Obtém o produto com o ID fornecido
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Verifica se um tamanho foi enviado no formulário
    tamanho_id = request.POST.get('size')
    if not tamanho_id:
        messages.error(request, 'É necessário selecionar um tamanho.')
        return redirect('store')

    # Verifica se o tamanho é válido para o produto
    tamanho_objeto = get_object_or_404(ProdutoTamanho, id=tamanho_id, produto=produto)

    # Define a chave do carrinho baseada no ID do produto e do tamanho
    chave_carrinho = f"{produto_id}"

    # Verifica se o produto já está no carrinho
    if chave_carrinho in cart:
        # Incrementa a quantidade do produto
        cart[chave_carrinho]['quantidade'] += 1
    else:
        # Adiciona o produto ao carrinho
        cart[chave_carrinho] = {
            'nome': produto.nome,
            'tamanho': tamanho_id,
            'size': tamanho_objeto.tamanho,
            'quantidade': 1,
            'sku': produto.sku,
            'preco': str(produto.preco),
            'categoria': produto.categoria
        }

    # Verifica se há estoque disponível para adicionar o item
    if cart[chave_carrinho]['quantidade'] > tamanho_objeto.stock_por_tamanho:
        messages.error(request, f'Não é possível adicionar a quantidade desejada ao carrinho. Estoque disponível para o tamanho {tamanho_objeto.tamanho}: {tamanho_objeto.stock_por_tamanho}.')
        return redirect('store')

    # Atualiza o carrinho na sessão
    request.session['carrinho'] = cart

    # Deduz o estoque correspondente ao tamanho selecionado
    tamanho_objeto.stock_por_tamanho -= 1
    tamanho_objeto.save()

    # Mensagem de sucesso
    messages.success(request, f'1 peça de {produto.nome} ({tamanho_objeto.tamanho}) adicionada ao carrinho.')

    # Redireciona para a página do carrinho
    return redirect('store')



def adicionar_roupa_dentro_carrinho(request, produto_id):
    # Obtém o carrinho da sessão
    cart = request.session.get('carrinho', {})
    
    # Obtém o produto com o ID fornecido
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Define a chave do carrinho baseada no ID do produto
    chave_carrinho = f"{produto_id}"

    # Verifica se o produto já está no carrinho
    if chave_carrinho in cart:
        # Incrementa a quantidade do produto
        cart[chave_carrinho]['quantidade'] += 1
    else:
        # Adiciona o produto ao carrinho
        cart[chave_carrinho] = {
            'nome': produto.nome,
            'tamanho': cart[chave_carrinho]['tamanho'],  # Utiliza o tamanho presente na cookie
            'size': cart[chave_carrinho]['size'],        # Utiliza o tamanho presente na cookie
            'quantidade': 1,
            'sku': produto.sku,
            'preco': str(produto.preco),
            'categoria': produto.categoria
        }

    # Verifica se há estoque disponível para adicionar o item
    tamanho_id = cart[chave_carrinho]['tamanho']
    tamanho_objeto = get_object_or_404(ProdutoTamanho, id=tamanho_id, produto=produto)
    if cart[chave_carrinho]['quantidade'] > tamanho_objeto.stock_por_tamanho:
        messages.error(request, f'Não é possível adicionar a quantidade desejada ao carrinho. Estoque disponível para o tamanho {tamanho_objeto.tamanho}: {tamanho_objeto.stock_por_tamanho}.')
        return redirect('carrinho')

    # Atualiza o carrinho na sessão
    request.session['carrinho'] = cart

    # Deduz o estoque correspondente ao tamanho selecionado
    tamanho_objeto.stock_por_tamanho -= 1
    tamanho_objeto.save()

    # Mensagem de sucesso
    messages.success(request, f'1 peça de {produto.nome} ({tamanho_objeto.tamanho}) adicionada ao carrinho.')

    # Redireciona para a página do carrinho
    return redirect('carrinho')




def remover_do_carrinho(request, produto_id):
    # Obtém o carrinho da sessão
    cart = request.session.get('carrinho', {})
    
    # Obtém o ID do produto como string
    product_id_str = str(produto_id)
    
    # Verifica se o produto está no carrinho
    if product_id_str in cart:
        # Decrementa a quantidade do item no carrinho
        if cart[product_id_str]['quantidade'] > 1:
            cart[product_id_str]['quantidade'] -= 1
        else:
            # Remove o item do carrinho se a quantidade for 1 ou menos
            del cart[product_id_str]

        # Atualiza o carrinho na sessão
        request.session['carrinho'] = cart

        # Incrementa o estoque do produto
        produto = get_object_or_404(Produto, id=produto_id)
        produto.stock += 1
        produto.save()

    return HttpResponseRedirect(reverse('carrinho'))


def remover_roupa_do_carrinho(request, produto_id):
    cart = request.session.get('carrinho', {})
    produto = get_object_or_404(Produto, id=produto_id)
    
    # Define a chave do carrinho baseada no ID do produto
    chave_carrinho = f"{produto_id}"

    # Verifica se o produto está presente no carrinho
    if chave_carrinho in cart:
        # Verifica se há mais de 1 item do produto no carrinho
        if cart[chave_carrinho]['quantidade'] > 1:
            # Remove 1 da quantidade do produto no carrinho
            cart[chave_carrinho]['quantidade'] -= 1
        else:
            # Remove completamente o produto do carrinho se a quantidade for 1
            del cart[chave_carrinho]
        
        # Atualiza a sessão do carrinho
        request.session['carrinho'] = cart

        # Adiciona 1 ao estoque do produto
        tamanho_id = cart[chave_carrinho].get('tamanho')
        if tamanho_id:
            produto_tamanho = get_object_or_404(ProdutoTamanho, id=tamanho_id)
            produto_tamanho.stock_por_tamanho += 1
            produto_tamanho.save()

        # Mensagem de sucesso
        messages.success(request, f'1 piece, {produto.nome}, was removed from the cart.')
    else:
        messages.error(request, 'The product is not in the cart')
        
    # Redireciona para a página do carrinho
    return redirect('carrinho')





def carrinho(request):
    countrys = country.objects.all()
    cart = request.session.get('carrinho', {})
    itens_carrinho = []
    total = 0
    print(cart)
    for produto_id, info_produto in cart.items():
        produto = get_object_or_404(Produto, id=produto_id)
        tamanho_id = info_produto.get('tamanho')
        tamanho = None
        tamanho_nome = 'unico'  # Definindo como 'unico' por padrão

        if tamanho_id != 'unico':
            tamanho = get_object_or_404(ProdutoTamanho, id=tamanho_id)
            tamanho_nome = tamanho.tamanho

        quantidade = info_produto['quantidade']  

        if quantidade > produto.stock:
            #messages.error(request, f'Not enough stock of {produto.nome}. Available: {produto.stock}.')
            # Adiciona ao carrinho mas marca como inativo e não adiciona ao total
            itens_carrinho.append({
                'produto_id': produto.id,
                'produto': produto,
                'quantidade': quantidade,
                'subtotal': produto.preco * min(quantidade, produto.stock),
                'tamanho': tamanho,
                'tamanho_nome': tamanho_nome,
                'ativo': True,  
            })
        else:
            subtotal = produto.preco * quantidade
            total += subtotal  
            itens_carrinho.append({
                'produto_id': produto.id,
                'produto': produto,
                'quantidade': quantidade,
                'subtotal': subtotal,
                'tamanho': tamanho,
                'tamanho_nome': tamanho_nome,
                'ativo': True,
            })

    return render(request, 'store/dados_encomenda.html', {
        'itens_carrinho': itens_carrinho,
        'total': total,
        'countrys': countrys,
    })


    
    
def produto(request, produto_id):
    # Utiliza get_object_or_404 para tentar obter o produto correspondente ao ID.
    # Caso não exista, retorna uma página 404 automaticamente.
    produto = get_object_or_404(Produto, id=produto_id)
    tamanhos_disponiveis = ProdutoTamanho.objects.filter(produto_id=produto,stock_por_tamanho__gt=0)
    
    context={
        'produto': produto,
        'tamanhos_disponiveis': tamanhos_disponiveis
    }
    # Passa o produto obtido para o template.
    return render(request, 'produto.html', context)

def pessoa_encomenda(request):
    return render(request,'store/pessoa_encomenda.html')

def payout(request):
    return render(request,'store/payment.html')


def remover_dentro_carrinho(request, produto_id):
    # Acessa o carrinho armazenado na sessão
    carrinho = request.session.get('carrinho', {})

    # Converte o produto_id para string, pois as chaves do dicionário estão em string
    product_id_str = str(produto_id)

    # Verifica se o produto existe no carrinho e o remove
    if product_id_str in carrinho:
        del carrinho[product_id_str]
        request.session['carrinho'] = carrinho  # Salva o carrinho atualizado na sessão

    # Redireciona para a página do carrinho ou para outra página de sua escolha
    return redirect('carrinho')




import csv
from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from Void.models import Evento, Produto, ProdutoTamanho, country
from django.views.decorators.http import require_POST
from django.contrib import messages
import openpyxl
from django.core.files.base import ContentFile
from .forms import EventoForm, ProdutoForm, ProdutoTamanhoForm,UploadExcelForm, paisesForm 
from django.contrib.auth import authenticate, login,logout
from django.urls import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.core.paginator import Paginator
from django.db.models import Sum


def lista_produtos_tamanhos(request):
    
    produtos_list = Produto.objects.all()
    tamanhos_list = ProdutoTamanho.objects.all()
    for produto in produtos_list:
        tamanhos_sum = ProdutoTamanho.objects.filter(produto_id=produto).aggregate(total=Sum('stock_por_tamanho'))
        produto.stock = tamanhos_sum['total'] if tamanhos_sum['total'] else produto.stock
        
    paginator_produtos = Paginator(produtos_list, 10)  # Mostra 10 produtos por página
    page_number_produtos = request.GET.get('page')
    page_obj_produtos = paginator_produtos.get_page(page_number_produtos)

    if request.method == 'POST':
        
        form = UploadExcelForm(request.POST, request.FILES)
        
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(filename=ContentFile(excel_file.read()))
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                produto = Produto(
                    nome=row[0],
                    descricao=row[1],
                    preco=row[2],
                    stock=row[3],
                    sku=row[4],
                    categoria=row[5],
                    cor=row[6],
                    tamanho=row[7],
                )

                imagem_url = row[8] if len(row) > 9 else None# Ajuste o índice conforme a sua planilha
                if imagem_url:
                    try:
                        response = request.get(imagem_url)
                        img_temp = ContentFile(response.content)
                        produto.imagem.save(f"{row[5]}_image.jpg", img_temp)
                    except Exception as e:
                        print(f"Erro ao baixar a imagem de {imagem_url}: {e}")

                produto.save()

            return redirect('lista_produtos')
    else:
        form = UploadExcelForm()
    return render(request, 'Adm/Produto/lista_produtos.html', {'page_obj': page_obj_produtos,'form':form,'tamanhos': tamanhos_list})



def adicionar_produto(request):
    if request.method == 'POST':
        form = ProdutoForm(request.POST, request.FILES)  # Inclua request.FILES aqui
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Produto adicionado com sucesso!')
            return HttpResponseRedirect(reverse('lista_produtos'))
    else:
        form = ProdutoForm()
    return render(request, 'Adm/Produto/form_produto.html', {'form': form})



def editar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == 'POST':
        form = ProdutoForm(request.POST, request.FILES, instance=produto)  # Inclua request.FILES aqui
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Produto editado com sucesso!')
            return redirect('lista_produtos')
    else:
        form = ProdutoForm(instance=produto)
    return render(request, 'Adm/Produto/form_produto.html', {'form': form})

def eliminar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)
    if request.method == "POST":
        produto.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'Produto eliminado com sucesso!')
        return redirect('lista_produtos')
    return render(request, 'Adm/Produto/confirmar_eliminar.html', {'produto': produto})

@require_POST
def eliminar_selecionados(request):
    # Obtém a lista de IDs dos produtos selecionados
    produtos_ids = request.POST.getlist('produto_id')
    
    # Elimina os produtos selecionados
    Produto.objects.filter(id__in=produtos_ids).delete()
    
    # Redireciona para a página de listagem de produtos
    return redirect('lista_produtos')

from django.shortcuts import redirect, render
from .forms import ProdutoTamanhoForm

def adicionar_tamanho(request):
    if request.method == 'POST':
        form = ProdutoTamanhoForm(request.POST)
        if form.is_valid():
            tamanho = form.save(commit=False)
            tamanho.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Tamanho adicionado com sucesso!')
            return HttpResponseRedirect(reverse('lista_produtos'))
    else:
        form = ProdutoTamanhoForm()
    return render(request, 'Adm/Produto/tamanhoForm.html', {'form': form})

def editar_tamanho(request, id):
    tamanho = get_object_or_404(ProdutoTamanho, id=id)
    if request.method == "POST":
        form = ProdutoTamanhoForm(request.POST, instance=tamanho)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Tamanho editado com sucesso!')
            return HttpResponseRedirect(reverse('lista_produtos'))
    else:
        form = ProdutoTamanhoForm(instance=tamanho)
    
    return render(request, 'Adm/Produto/editar_tamanho.html', {'form': form})

def excluir_tamanho(request, id):
    tamanho = get_object_or_404(ProdutoTamanho, id=id)
    if request.method == "POST":
        tamanho.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'Tamanho eliminado com sucesso!')
        return HttpResponseRedirect(reverse('lista_produtos'))
    
    return render(request, 'Adm/Produto/confirmar_exclusao_tamanho.html', {'tamanho': tamanho})

def login_view(request):
    # Se o user já está logado, redirecioná-lo.
    if request.user.is_authenticated:
        return redirect('admin')
    if request.method == 'POST':
        # Pegar as informações do formulário.
        username = request.POST.get('email')
        password = request.POST.get('password')

        # Autenticar o usuário.
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request,"Login bem sucedido") 
            return HttpResponseRedirect(reverse('admin'))
        else:
            # Se a autenticação falhou, retorne algum erro.
            messages.error(request,'Email ou password incorretos.')
            return HttpResponseRedirect(reverse('login'))
        # Método GET, renderize o formulário de login.
    return render(request, 'Adm/login.html')
    

def adm(request):
    
    return render(request,'Adm/adm.html')

def logout_view(request):
    logout(request)
    # Adiciona uma mensagem de sucesso
    messages.success(request, 'Logout efetuado com sucesso!')
    return HttpResponseRedirect(reverse('home'))

def exportar_produtos_csv(request):
    # Criar uma resposta do tipo arquivo
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="produtos.csv"'

    # Criar um escritor CSV usando a resposta como arquivo
    writer = csv.writer(response,delimiter=';')

    # Escrever o cabeçalho do CSV
    writer.writerow(['ID', 'Nome', 'Descrição', 'Preço', 'Estoque', 'SKU', 'Categoria', 'Cor', 'Tamanho'])

    # Obter os produtos da base de dados
    produtos = Produto.objects.all()

    # Escrever os dados de cada produto
    for produto in produtos:
        writer.writerow([produto.id, produto.nome, produto.descricao, produto.preco, produto.stock, produto.sku, produto.categoria, produto.cor, produto.tamanho])

    # Retornar a resposta que agora contém o arquivo CSV
    return response

def paises_adicionar(request):
    if request.method == 'POST':
        form = paisesForm(request.POST)  # Inclua request.FILES aqui
        if form.is_valid():
            form.save()
           # Adiciona uma mensagem de sucesso
            messages.success(request, 'País adicionado com sucesso!')
            return HttpResponseRedirect(reverse('listar_paises'))
    else:
        form = paisesForm()
    return render(request, 'Adm/Paises/paisesForm.html', {'form': form})

def listar_paises(request):
    paises = country.objects.all()
    if request.method == 'POST':
        excel_file = request.FILES["excel_file"]
        
        # Usa a biblioteca openpyxl para abrir o arquivo Excel
        wb = openpyxl.load_workbook(excel_file)
        # Seleciona a primeira planilha do arquivo
        worksheet = wb["Sheet1"]
        
        # Itera sobre as linhas da planilha
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            _, created = country.objects.update_or_create(
                name=row[0],
                acronimo=row[1],
                shipping=row[2],
            )
        return HttpResponseRedirect(reverse('listar_paises'))
    return render(request, 'Adm/Paises/paisesLista.html', {'paises': paises})

def editar_pais(request, id):
    pais = get_object_or_404(country, id=id)
    if request.method == "POST":
        form = paisesForm(request.POST, instance=pais)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'País editado com sucesso!')
            return HttpResponseRedirect(reverse('listar_paises'))
    else:
        form = paisesForm(instance=pais)
    return render(request, 'Adm/Paises/paisesForm.html', {'form': form})

# Exclui país
def excluir_pais(request, id):
    pais = get_object_or_404(country, id=id)
    if request.method == "POST":
        pais.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'País eliminado com sucesso!')
        return HttpResponseRedirect(reverse('listar_paises'))
    return render(request, 'Adm/Paises/excluirPais.html', {'pais': pais})




def lista_eventos(request):
    eventos = Evento.objects.all()
    return render(request, 'Adm/eventos/lista_eventos.html', {'eventos': eventos})

def evento_add(request):
    if request.method == "POST":
        form = EventoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Ritual adicionado com sucesso!')
            return HttpResponseRedirect(reverse('lista_eventos'))
    else:
        form = EventoForm()
    return render(request, 'Adm/Eventos/form.html', {'form': form})

def evento_edit(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        form = EventoForm(request.POST, request.FILES, instance=evento)
        if form.is_valid():
            form.save()
            # Adiciona uma mensagem de sucesso
            messages.success(request, 'Ritual editado com sucesso!')
            return HttpResponseRedirect(reverse('lista_eventos'))
    else:
        form = EventoForm(instance=evento)
    return render(request, 'Adm/Eventos/form.html', {'form': form})

def evento_delete(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        evento.delete()
        # Adiciona uma mensagem de sucesso
        messages.success(request, 'Ritual eliminado com sucesso!')
        return HttpResponseRedirect(reverse('lista_eventos'))
    return render(request, 'Adm/Eventos/delete.html', {'evento': evento})

def exportar_eventos_para_csv(request):
    # Cria a resposta com o tipo de conteúdo correto
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="eventos.csv"'

    # Cria um objeto writer para escrever no arquivo CSV
    writer = csv.writer(response)

    # Escreve o cabeçalho do CSV
    writer.writerow(['ID', 'Título', 'Data', 'Localização', 'Hiperligação', 'Caminho da Foto'])

    # Escreve as linhas de dados dos eventos
    eventos = Evento.objects.all()
    for evento in eventos:
        writer.writerow([
            evento.id, 
            evento.titulo, 
            evento.data, 
            evento.localizacao, 
            evento.hiperligacao, 
            request.build_absolute_uri(evento.foto.url) if evento.foto else None,
        ])

    return response

def lista_pagamentos(request):
    pagamentos = Pagamento.objects.all()
    return render(request, 'Adm/Pagamentos/pagamentos.html', {'pagamentos': pagamentos})

def detalhes_pagamento(request, pagamento_id):
    pagamento = get_object_or_404(Pagamento, pk=pagamento_id)
    itens_do_carrinho = ItemDoCarrinho.objects.filter(pagamento=pagamento)
    return render(request, 'Adm/Pagamentos/detalhes_pagamento.html', {'pagamento': pagamento, 'itens_do_carrinho': itens_do_carrinho})

import logging

# Configurar o logger
logger = logging.getLogger(__name__)


@require_POST
def create_payment(request):
    try:
        # Obtenha os itens do carrinho do cookie
        cart = request.session.get('carrinho', {})

        # Verifique se todos os produtos estão em estoque
        for produto_id, info_produto in cart.items():
            produto = get_object_or_404(Produto, id=produto_id)
            quantidade = info_produto['quantidade']
            #if produto.stock < quantidade:
                #messages.error(request, f"O produto {produto.nome} está fora de estoque.")
             #   return HttpResponseRedirect(reverse('carrinho'))  # Redirecione de volta para a página do carrinho


        acronimo_pag = request.POST.get('pais')

        # Obtém o custo de envio baseado no país
        shippin = get_object_or_404(country, acronimo=acronimo_pag)
        shipping_cost = Decimal(shippin.shipping).quantize(Decimal('.01'))  # Assegura 2 casas decimais

        # Crie uma lista de itens para o PayPal
        paypal_items = []
        total_amount = Decimal(0)
        subtotal = Decimal(0)

        for produto_id, info_produto in cart.items():
            produto = get_object_or_404(Produto, id=produto_id)
            quantidade = info_produto['quantidade']
            item_price = Decimal(produto.preco).quantize(Decimal('.01'))  # Assegura 2 casas decimais
            paypal_items.append({
                "name": produto.nome,
                "sku": produto.sku,
                "price": str(item_price),  # Convertido para string
                "currency": "EUR",
                "quantity": quantidade,
            })

            subtotal += item_price * quantidade  # Subtotal dos itens

        # Adiciona o custo de envio ao total
        total_amount = (subtotal + shipping_cost).quantize(Decimal('.01'))  # Total final assegurado

        print(f"Subtotal dos Itens: {total_amount - shipping_cost}")
        print(f"Custo de Envio: {shipping_cost}")
        print(f"Total Enviado na Transação: {total_amount}")

        
        # Preparação para a criação do pagamento
        payment = Payment({
            "intent": "sale",
            "payer": {
                "payment_method": "paypal",
            },
            "redirect_urls": {
                "return_url": request.build_absolute_uri(reverse("execute_payment")),
                "cancel_url": request.build_absolute_uri(reverse("payment_cancelled")),
            },
            "transactions": [{
                "item_list": {
                    "items": paypal_items,
                },
                "amount": {
                    "total": str(total_amount),
                    "currency": "EUR",
                    "details":{
                        "subtotal":str(subtotal),
                        "shipping":str(shipping_cost),
                    }
                },
                "description": "Descrição do pedido.",
            }]
        })

        if payment.create():
            for link in payment.links:
                if link.rel == "approval_url":
                    approval_url = str(link.href)
                    return HttpResponseRedirect(approval_url)
            return HttpResponse("Não foi possível processar o pagamento.", status=500)
        else:
            # Supondo que você tenha um logger configurado
            logger.error("Erro ao criar o pagamento. Detalhes: %s" % payment.error)
            return HttpResponse("Erro ao criar o pagamento.", status=500)
    except Exception as e:
        logger.error("Erro ao criar o pagamento: %s" % str(e))
        return HttpResponse("Erro ao criar o pagamento.", status=500)
    

def execute_payment(request):
    try:
        payment_id = request.GET.get('paymentId')
        payer_id = request.GET.get('PayerID')

        payment = Payment.find(payment_id)

        # Verifica se o pagamento foi encontrado e se a execução foi bem-sucedida
        if payment and payment.execute({"payer_id": payer_id}):
            messages.success(request, 'Payment executed with success!')
            return HttpResponseRedirect(reverse('payment_suc', args=[payment_id]))  # Retorna uma resposta HTTP de sucesso
        else:
            # Falha ao executar o pagamento
            messages.error(request, 'Failed to execute payment.')
            return HttpResponse("Falha ao executar o pagamento.", status=400)
    except Exception as e:
        logger.error("Erro ao executar o pagamento: %s" % str(e))
        messages.error(request, 'Failed to execute payment.')
        return HttpResponse("Erro ao executar o pagamento.", status=500)


def payment_success(request, payment_id):
    try:
        payment = Payment.find(payment_id)

        if payment:
            payer_info = payment.payer.payer_info
            shipping_address = payer_info.shipping_address

            # Verifica se as informações de envio estão disponíveis
            if shipping_address:
                # Crie um novo objeto Pagamento
                pagamento = Pagamento.objects.create(
                    order_id=payment_id,
                    email=payer_info.email,
                    endereco=shipping_address.line1,
                    codigo_postal=shipping_address.postal_code,
                    cidade=shipping_address.city,
                    total=Decimal(payment.transactions[0].amount.total),
                )

                cart = request.session.get('carrinho', {})
                print(cart)
                for produto_id, info_produto in cart.items():
                    produto = get_object_or_404(Produto, id=produto_id)


                    # # Verifica se o produto pertence à categoria "roupa"
                    # if produto.categoria == 'roupa':
                    #     id_tamanho = info_produto['tamanho']
                    #     print(id_tamanho)
                    #     tamanho_produto = get_object_or_404(ProdutoTamanho, produto=produto, id=id_tamanho)
                    quantidade_comprada = info_produto['quantidade']
                    #     tamanho_produto.stock_por_tamanho -= quantidade_comprada
                    #     tamanho_produto.save()
                    # else:
                    #     quantidade_comprada = info_produto['quantidade']
                    #     produto.stock -= quantidade_comprada
                    #     produto.save()

                    ItemDoCarrinho.objects.create(
                        pagamento=pagamento,
                        produto_id=produto_id,
                        nome=produto.nome,
                        preco=produto.preco,
                        quantidade=quantidade_comprada  # Use a quantidade_comprada do loop atual
                    )

                # Limpe o carrinho após o pagamento ser concluído
                del request.session['carrinho']
                return render(request, 'Payment/sucess.html')
            else:
                # As informações de envio não estão disponíveis
                messages.error(request, 'Failed to retrieve shipping address.')
                return HttpResponse("Falha ao obter o endereço de envio.", status=400)
        else:
            # O objeto de pagamento não foi encontrado
            messages.error(request, 'Failed to retrieve payment information.')
            return HttpResponse("Falha ao obter informações de pagamento.", status=400)
    except Exception as e:
        logger.error("Erro ao processar o pagamento: %s" % str(e))
        messages.error(request, 'Failed to process payment.')
        return HttpResponse("Erro ao processar o pagamento.", status=500)




def payment_cancelled(request):
    try:
        # Obtém o carrinho da sessão
        cart = request.session.get('carrinho', {})

        # Restaura o estoque dos produtos que estavam no carrinho
        for produto_id, info_produto in cart.items():
            produto = get_object_or_404(Produto, id=produto_id)

            # Verifica se o produto pertence à categoria "roupa"
            if produto.categoria == 'roupa':
                id_tamanho = info_produto['tamanho']
                tamanho_produto = get_object_or_404(ProdutoTamanho, produto=produto, id=id_tamanho)
                quantidade_comprada = info_produto['quantidade']
                tamanho_produto.stock_por_tamanho += quantidade_comprada
                tamanho_produto.save()
            else:
                quantidade_comprada = info_produto['quantidade']
                produto.stock += quantidade_comprada
                produto.save()

        # Limpa o carrinho após o cancelamento do pagamento
        del request.session['carrinho']

        # Retorna uma mensagem indicando que o pagamento foi cancelado pelo usuário
        return HttpResponse("Pagamento cancelado pelo usuário.")
    except Exception as e:
        logger.error("Erro ao lidar com o cancelamento do pagamento: %s" % str(e))
        messages.error(request, 'Erro ao lidar com o cancelamento do pagamento.')
        return HttpResponse("Erro ao lidar com o cancelamento do pagamento.", status=500)
